import sys
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter  # Proper column conversion

if __name__ == "__main__":
    # Step 1: Get log date from sys.argv
    sheet_name = sys.argv[1]  # Extract log date

    # Step 2: Read the Excel file buffer from stdin
    file_bytes = sys.stdin.buffer.read()

# Step 3: Load the Excel file into memory using openpyxl
wb = load_workbook(BytesIO(file_bytes))

# Step 4: Access the specific sheet by name
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
ws = wb[sheet_name]

def create_line_charts(ws, start_data_row=4, end_data_row=16, start_chart_cell="A25", row_spacing=15, col_spacing=14):
    """
    Create multiple line charts in a 3-row by 4-column layout with proper alignment.
    
    :param ws: The worksheet object
    :param start_data_row: First data row containing the titles
    :param end_data_row: Last data row containing the titles
    :param start_chart_cell: Starting cell for chart placement (e.g., "A25")
    :param row_spacing: Number of rows between each chart vertically
    :param col_spacing: Number of columns between each chart horizontally
    """

    # Define the layout (3 rows x 4 columns)
    charts_per_row = 4  # 4 columns per row

    # Extract starting chart position
    start_col = ord(start_chart_cell[0]) - ord('A') + 1  # Convert 'A' -> 1
    start_row = int(''.join(filter(str.isdigit, start_chart_cell)))  # Extract row number (25)

    # Create a reference for the x-axis categories from S3:X3
    x_ref = Reference(ws, min_col=19, max_col=24, min_row=3, max_row=3)

    # Track chart index
    chart_index = 0

    # Loop over each row with a chart title in column A
    for r in range(start_data_row, end_data_row + 1):
        chart_title = ws.cell(row=r, column=1).value

        # Create a new LineChart
        chart = LineChart()
        chart.title = chart_title
        chart.style = 2
        chart.x_axis.number_format = "@"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.set_categories(x_ref)

        # Define Y-axis data ranges for each engine using the current row:
        eng_refs = [
            Reference(ws, min_col=7, max_col=12, min_row=r, max_row=r),  # Engine 1 (G:L)
            Reference(ws, min_col=13, max_col=18, min_row=r, max_row=r),  # Engine 2 (M:R)
            Reference(ws, min_col=19, max_col=24, min_row=r, max_row=r),  # Engine 3 (S:X)
            Reference(ws, min_col=25, max_col=30, min_row=r, max_row=r),  # Engine 4 (Y:AD)
            Reference(ws, min_col=31, max_col=36, min_row=r, max_row=r),  # Engine 5 (AE:AJ)
        ]

        # Add series for each engine
        for ref, title in zip(eng_refs, ["Engine 1", "Engine 2", "Engine 3", "Engine 4", "Engine 5"]):
            series = Series(ref, title=title)
            series.graphicalProperties.line.smooth = False
            chart.series.append(series)

        # Compute grid position
        row_pos = chart_index // charts_per_row  # Determine row position (0,1,2)
        col_pos = chart_index % charts_per_row   # Determine column position (0,1,2,3)

        # Compute anchor cell using `get_column_letter`
        anchor_row = start_row + (row_pos * row_spacing)  # Adjust by row_spacing
        anchor_col = start_col + (col_pos * col_spacing)  # Adjust by col_spacing
        anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"  # Convert to Excel column letter format

        # Add chart to the worksheet
        ws.add_chart(chart, anchor_cell)

        # Increment chart index
        chart_index += 1

# Step 4: Process the Excel file (add charts)
create_line_charts(ws, start_data_row=4, end_data_row=16, start_chart_cell="A25", row_spacing=15, col_spacing=14)

# Step 5: Save the modified file to an in-memory buffer
output_buffer = BytesIO()
wb.save(output_buffer)

# Step 6: Write the modified buffer back to stdout
sys.stdout.buffer.write(output_buffer.getvalue())
